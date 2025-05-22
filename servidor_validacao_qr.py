from flask import Flask, request, send_file
import os
import json
from io import BytesIO
from openpyxl import load_workbook

app = Flask(__name__)

ARQUIVO_USADOS = "usados.json"
MODELO_XLSX = "modelo.xlsx"

# Cria o arquivo JSON se não existir
if not os.path.exists(ARQUIVO_USADOS):
    with open(ARQUIVO_USADOS, "w") as f:
        json.dump([], f)

@app.route("/")
def home():
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Scanner QR Code</title>
        <style>
            body { font-family: Arial, sans-serif; background: #f7f7f7; text-align: center; padding: 20px; margin: 0; }
            #reader { width: 100%; max-width: 400px; margin: 20px auto; }
            button { font-size: 18px; padding: 12px 24px; margin-top: 30px; cursor: pointer; background: #007bff; color: white; border: none; border-radius: 8px; }
            .box { margin: 20px auto; max-width: 90%; font-size: 20px; color: #333; }
            a.usados-link { display: inline-block; margin-top: 20px; font-size: 18px; color: #007bff; text-decoration: none; }
            a.usados-link:hover { text-decoration: underline; }
        </style>
        <script src="https://unpkg.com/html5-qrcode"></script>
    </head>
    <body>
        <div class="box" id="mensagem">Clique em "Escanear QR Code" para começar</div>
        <button id="btnScan">Escanear QR Code</button>
        <div id="reader" style="display:none;"></div>
        <br>
        <a href="/usados" class="usados-link">Ver QR Codes escaneados</a>

        <script>
            const btnScan = document.getElementById('btnScan');
            const reader = document.getElementById('reader');
            const mensagemBox = document.getElementById('mensagem');
            let scanner;

            btnScan.addEventListener('click', () => {
                btnScan.style.display = 'none';
                reader.style.display = 'block';

                scanner = new Html5Qrcode("reader");

                Html5Qrcode.getCameras().then(devices => {
                    if (devices && devices.length) {
                        const backCamera = devices.find(device => device.label.toLowerCase().includes('back')) || devices[0];

                        scanner.start(
                            backCamera.id,
                            { fps: 10, qrbox: 250 },
                            (decodedText) => {
                                scanner.stop().then(() => {
                                    scanner.clear();
                                    reader.style.display = 'none';
                                    btnScan.style.display = 'block';

                                    mensagemBox.innerHTML = `Validando código...`;

                                    fetch('/validar', {
                                        method: 'POST',
                                        headers: { 'Content-Type': 'application/json' },
                                        body: JSON.stringify({ codigo: decodedText })
                                    })
                                    .then(response => response.text())
                                    .then(html => {
                                        mensagemBox.innerHTML = html;
                                        scanner = null;
                                    })
                                    .catch(err => {
                                        mensagemBox.innerHTML = "Erro ao validar código.";
                                    });
                                }).catch(err => {
                                    console.error('Erro ao parar o scanner', err);
                                });
                            },
                            errorMessage => {}
                        ).catch(err => {
                            mensagemBox.innerHTML = "Não foi possível iniciar a câmera.";
                            btnScan.style.display = 'block';
                        });
                    } else {
                        mensagemBox.innerHTML = "Nenhuma câmera encontrada.";
                        btnScan.style.display = 'block';
                    }
                }).catch(err => {
                    mensagemBox.innerHTML = "Erro ao acessar câmeras: " + err;
                    btnScan.style.display = 'block';
                });
            });
        </script>
    </body>
    </html>
    """

@app.route("/validar", methods=["POST"])
def validar_qrcode():
    data = request.get_json()
    codigo = data.get("codigo", "")

    try:
        dados = json.loads(codigo)

        with open(ARQUIVO_USADOS, "r") as f:
            usados = json.load(f)

        if codigo in usados:
            mensagem = "<div style='color:#ff4d4d; font-weight:bold; font-size:24px;'>⚠️ QRCODE JÁ USADO!</div><small>Este código já foi validado anteriormente.</small>"
            cor = "#ffdddd"
        else:
            mensagem = "<div style='color:#4CAF50; font-weight:bold; font-size:24px;'>✅ QRCODE VÁLIDO!</div><small>Bem-vindo ao evento 🎉</small>"
            usados.append(codigo)
            with open(ARQUIVO_USADOS, "w") as f:
                json.dump(usados, f, indent=4)
            cor = "#ddffdd"

        tipo = dados.get("tipo", "").lower()
        if tipo == "titular":
            cabecalho = "<h3 style='color:#007bff;'>🎫 Titular</h3>"
        elif tipo == "acompanhante":
            cabecalho = "<h3 style='color:#555;'>👥 Acompanhante</h3>"
        else:
            cabecalho = ""

        detalhes = "<ul style='list-style:none; padding-left:0; font-size:18px; text-align:left;'>"
        for chave, valor in dados.items():
            detalhes += f"<li><strong>{chave.replace('_', ' ').title()}:</strong> {valor}</li>"
        detalhes += "</ul>"

        html = f"""
        <div style="background:{cor};padding:20px;border-radius:10px; max-width:500px; margin:20px auto; color:#333; font-family: Arial, sans-serif;">
            {mensagem}
            {cabecalho}
            {detalhes}
        </div>
        """
    except json.JSONDecodeError:
        html = "<div style='color:red;'>❌ Erro: Código inválido ou malformado.</div>"

    return html

@app.route("/usados")
def listar_usados():
    if os.path.exists(ARQUIVO_USADOS):
        with open(ARQUIVO_USADOS, "r") as f:
            usados = json.load(f)
    else:
        usados = []

    html = """
    <html>
    <head>
        <title>QR Codes Usados</title>
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #eef1f5; margin: 0; padding: 0; }
            .container { max-width: 800px; margin: 0 auto; padding: 40px 20px; }
            h1 { text-align: center; margin-bottom: 30px; color: #333; }
            .qrcode-list { display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 20px; list-style: none; padding: 0; }
            .qrcode-card { background: white; border-radius: 12px; padding: 20px; box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1); font-size: 16px; word-break: break-word; display: flex; align-items: center; gap: 12px; transition: transform 0.2s; }
            .qrcode-card:hover { transform: scale(1.02); }
            .icon { font-size: 20px; color: #4CAF50; }
            .back-link { display: block; margin: 30px auto 0; text-align: center; font-size: 18px; color: #007bff; text-decoration: none; }
            .back-link:hover { text-decoration: underline; }
        </style>
    </head>
    <body>
        <div class="container">
            <h1>QR Codes Escaneados</h1>
            <ul class="qrcode-list">
    """

    for codigo in usados:
        html += f"""
        <li class="qrcode-card">
            <span class="icon">✅</span> {codigo}
        </li>
        """

    html += """
            </ul>
            <a class="back-link" href="/exportar_excel">⬇️ Baixar Excel</a>
            <a class="back-link" href="/">← Voltar para a página inicial</a>
        </div>
    </body>
    </html>
    """
    return html

@app.route("/exportar_excel")
def exportar_excel():
    if not os.path.exists(ARQUIVO_USADOS):
        return "Nenhum dado para exportar", 404
    if not os.path.exists(MODELO_XLSX):
        return "Modelo de planilha não encontrado", 404

    with open(ARQUIVO_USADOS, "r") as f:
        usados = json.load(f)

    try:
        wb = load_workbook(MODELO_XLSX)
        ws = wb.active

        for idx, item in enumerate(usados, start=2):
            dados = json.loads(item)
            ws.cell(row=idx, column=1, value=dados.get("nome", ""))
            ws.cell(row=idx, column=2, value=dados.get("tipo", ""))
            ws.cell(row=idx, column=3, value=dados.get("documento", ""))
            ws.cell(row=idx, column=4, value=dados.get("email", ""))
            ws.cell(row=idx, column=5, value=dados.get("telefone", ""))
    except Exception as e:
        return f"Erro ao exportar planilha: {e}", 500

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="qrcodes_exportados.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
